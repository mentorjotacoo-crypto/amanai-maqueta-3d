import openpyxl, re, sys
from openpyxl.styles import Font, PatternFill
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
SRC=r'i:\.shortcut-targets-by-id\1dODjY-KJP7NoYpWTtC5vuPhyIn84XUqk\RESERVAS DE AMANAI\ÁREAS URBANISMO AMANAI RADICADO DICIEMBRE 2025 (2).xlsx'
FINAL={
 1:{'T1':{1,11},'T2':{2,10,5,6}},
 2:{'T1':{1,10},'T2':{2,9,5,6}},
 3:{'T1':{1,7},'T2':{2,6,4,5}},
 4:{'T1':{1,8},'T2':{2,7,4,5}},
 5:{'T1':{1},'T2':{2,5}},
 12:{'T1':{1},'T2':{2,12}},
}
LBL={'T1':'ESQUINERO TIPO 1','T2':'ESQUINERO TIPO 2','R':'UNIFAMILIAR'}
t1f=PatternFill('solid',fgColor='F4CCCC'); t2f=PatternFill('solid',fgColor='FCE5CD'); nof=PatternFill(fill_type=None)
wb=openpyxl.load_workbook(SRC); ws=wb['ÁREAS FINAL JUL. 2026 (2)']
changed=0
for r in range(4, ws.max_row+1):
    mz=ws.cell(r,2).value; desc=ws.cell(r,3).value
    if mz is None or desc is None: continue
    m=re.match(r'LOTE\s+(\d+)',str(desc).upper().strip()); mzn=re.search(r'(\d+)',str(mz))
    if not (m and mzn): continue
    mzi=int(mzn.group(1)); n=int(m.group(1))
    if mzi not in FINAL: continue
    dd=FINAL[mzi]
    typ='T1' if n in dd['T1'] else ('T2' if n in dd['T2'] else 'R')
    cell=ws.cell(r,6)
    new=LBL[typ]
    if cell.value!=new: changed+=1
    cell.value=new
    cell.font=Font(name='Verdana',size=10)  # sin cursiva: ya no esta en revision
    cell.fill = t1f if typ=='T1' else (t2f if typ=='T2' else nof)
    # limpiar observacion de revision
    obs=ws.cell(r,7)
    if obs.value: obs.value=None
wb.save(SRC)
print('celdas cambiadas:', changed)
